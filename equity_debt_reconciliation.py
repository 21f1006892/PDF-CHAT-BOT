import os
import math
import subprocess
from copy import copy
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ===================== FILE PATHS =====================
SOURCE_FILE_XLS = "SEBI_Monthly_Portfolio 31 JAN 2026.xls"
SOURCE_FILE_XLSX = "SEBI_Monthly_Portfolio 31 JAN 2026.xlsx"
MASTER_FILE = "IN_MF_PORTFOLIO_DETAILS_ACROSS_SCHEMES.xlsx"
OUTPUT_FILE = "SEBI_Monthly_Portfolio_Equity_Debt_Reconciled.xlsx"

# IMPORTANT:
# In the uploaded master file, Total Market Value (Rs.) matches SEBI's
# "Market/Fair Value\n(Rs.in Lacs)" after dividing by 100000 because 1 lac = 100000.
MARKET_VALUE_DIVISOR = 100000

# ===================== SCHEME TYPE DICTIONARY FROM TESTTWO.PY =====================
scheme_dict = {'ABBSEIIF': 'EQUITY', 'ABSLBCF': 'EQUITY', 'ABSLCONF': 'EQUITY', 'ABSLESG': 'EQUITY',
'ABSLLDF': 'DEBT', 'ABSLMAAF': 'HYBRID', 'ABSLMCF': 'EQUITY', 'ABSLQF': 'EQUITY', 'ABSLSO': 'EQUITY',
'ABSLTNLF': 'EQUITY', 'ABSLUS03': 'FOF', 'ABSLUS10': 'FOF', 'ADVG': 'EQUITY', 'BANKETF': 'EQUITY', 'BBIF': 'DEBT',
'BBP': 'DEBT', 'BDB': 'DEBT', 'BDYP': 'EQUITY', 'BFL': 'DEBT', 'BFS': 'DEBT', 'BINFRA': 'EQUITY',
'BINTEQA': 'INTRNL', 'BMIDX50': 'EQUITY', 'BQIDX50': 'EQUITY', 'BSL95F': 'HYBRID', 'BSLAAMM': 'FOF',
'BSLADMM': 'FOF', 'BSLBBYW': 'EQUITY', 'BSLBKFS': 'EQUITY', 'BSLCBF': 'DEBT', 'BSLCM': 'DEBT', 'BSLDAAF': 'HYBRID',
'BSLEAF': 'HYBRID', 'BSLEQSF': 'HYBRID', 'BSLEQTY': 'EQUITY', 'BSLFEF': 'EQUITY', 'BSLFPAP': 'FOF', 'BSLFPCP': 'FOF',
'BSLFPPP': 'FOF', 'BSLIF': 'DEBT', 'BSLMFG': 'EQUITY', 'BSLMIFOF': 'FOF', 'BSLMTP': 'DEBT', 'BSLNMF': 'EQUITY',
'BSLONF': 'DEBT', 'BSLPHF': 'EQUITY', 'BSLR96': 'EQUITY', 'BSLRF30': 'HYBRID', 'BSLRF40': 'HYBRID',
'BSLRF50': 'HYBRID', 'BSLRF50P': 'DEBT', 'BSLSTF': 'DEBT', 'BSLTA1': 'EQUITY', 'BTOP100': 'EQUITY',
'C10YGETF': 'DEBT', 'CASH': 'DEBT', 'CBGETF': 'DEBT', 'CIGAPR26': 'DEBT', 'CIGAPR28': 'DEBT', 'CIGAPR29': 'DEBT',
'CIGAPR33': 'DEBT', 'CIGJUN27': 'DEBT', 'CIGPSA28': 'DEBT', 'CISJUN32': 'DEBT', 'COFAIETF': 'DEBT',
'CSFSD12': 'DEBT', 'CSFSD6': 'DEBT', 'CSFSI27': 'DEBT', 'CSNHFS26': 'DEBT', 'CSPAPA26': 'DEBT', 'CSPAPA27': 'DEBT',
'FTPTI': 'DEBT', 'FTPTJ': 'DEBT', 'FTPTQ': 'DEBT', 'FTPUB': 'DEBT', 'FTPUJ': 'DEBT', 'GENNEXT': 'EQUITY',
'GOLDETF': 'GOLD', 'GOLDFOF': 'FOF', 'INV': 'DEBT', 'MIDCAP': 'EQUITY', 'MIP25': 'HYBRID', 'MNC': 'EQUITY',
'N30MTETF': 'EQUITY', 'N30QTETF': 'EQUITY', 'NEWIF50': 'EQUITY', 'NHLTHETF': 'EQUITY', 'NIFTY': 'EQUITY',
'NIFTYETF': 'EQUITY', 'NIFYNX50': 'EQUITY', 'NINDDEF': 'EQUITY', 'NITETF': 'EQUITY', 'NMID150': 'EQUITY',
'NPSEETF': 'EQUITY', 'NSDAPR27': 'DEBT', 'NSDAQ100': 'FOF', 'NSDSEP27': 'DEBT', 'NSMALL50': 'EQUITY',
'NSPPBS26': 'DEBT', 'NXTIDX50': 'EQUITY', 'PLUS': 'DEBT', 'PSUEQ': 'EQUITY', 'PURE': 'EQUITY', 'SENSXETF': 'EQUITY',
'SILVRETF': 'SILVER', 'SILVRFOF': 'FOF', 'BSLGCF': 'OTH', 'BSLGRE': 'OTH'}

SEBI_COLUMNS = [
    "Name of the Instrument",
    "ISIN",
    "Industry^ / Rating",
    "Quantity",
    "Market/Fair Value\n(Rs.in Lacs)",
    "% to Net Assets",
]

MASTER_RENAME = {
    "Issuer Name": "Name of the Instrument",
    "Industry": "Industry^ / Rating",
    "Total Market Value (Rs.)": "Market/Fair Value\n(Rs.in Lacs)",
    "% to Net assests": "% to Net Assets",
}

COMPARE_COLS = [
    "Name of the Instrument",
    "Industry^ / Rating",
    "Quantity",
    "Market/Fair Value\n(Rs.in Lacs)",
    "% to Net Assets",
]

# ===================== DEBT-SPECIFIC COLUMNS =====================
# Keep debt separate from equity so equity reconciliation logic is not disturbed.
DEBT_SEBI_COLUMNS = [
    "Name of the Instrument",
    "ISIN",
    "Rating",
    "Quantity",
    "Market/Fair Value\n(Rs.in Lacs)",
    "% to Net Assets",
    "Yield",
]

DEBT_MASTER_RENAME = {
    "Security Name": "Name of the Instrument",
    "Total Market Value (Rs.)": "Market/Fair Value\n(Rs.in Lacs)",
    "% to Net assests": "% to Net Assets",
    "% to Net assets": "% to Net Assets",
}

DEBT_COMPARE_COLS = [
    "Name of the Instrument",
    "Rating",
    "Quantity",
    "Market/Fair Value\n(Rs.in Lacs)",
    "% to Net Assets",
    "Yield",
]

DEBT_MASTER_SECURITY_TYPES = [
    "Alternative Investment Funds (AIF)",
    "CERTIFICATE OF DEPOSIT",
    "COLLATERALISED BORROWING AND LENDIN",
    "COMMERCIAL PAPERS",
    "Cash Management Bills",
    "Fixed rates bonds - Corporate",
    "Fixed rates bonds - Government",
    "Fixed rates bonds - State Government",
    "Floating rates notes - Corporate",
    "Infra Investment Trust (INVITS)",
    "Interest Rate Swaps",
    "REITS",
    "REPO",
    "SECURITISED DEBT",
    "TREASURY BILLS",
    "Zero Coupon Bonds - Corporate",
]

HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")     # yellow
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")         # light red
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def resolve_source_xlsx():
    """
    Return an xlsx file path that openpyxl can edit.

    openpyxl cannot edit old .xls files directly.
    So this function first looks for an already-saved .xlsx file.
    If only .xls exists, it tries:
      1) Microsoft Excel COM conversion on Windows, if Excel + pywin32 are available.
      2) LibreOffice conversion, if LibreOffice is installed and available in PATH.

    Easiest manual fix if conversion fails:
      Open the SEBI .xls file in Excel -> File -> Save As -> .xlsx
      and keep the file name as SOURCE_FILE_XLSX below.
    """
    if os.path.exists(SOURCE_FILE_XLSX):
        return SOURCE_FILE_XLSX

    if not os.path.exists(SOURCE_FILE_XLS):
        raise FileNotFoundError(
            f"Could not find {SOURCE_FILE_XLSX} or {SOURCE_FILE_XLS} in this folder."
        )

    # Method 1: Windows + Microsoft Excel installed
    try:
        import win32com.client  # pip install pywin32

        src = os.path.abspath(SOURCE_FILE_XLS)
        dst = os.path.abspath(SOURCE_FILE_XLSX)

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_excel = excel.Workbooks.Open(src)
        # 51 = xlOpenXMLWorkbook = .xlsx
        wb_excel.SaveAs(dst, FileFormat=51)
        wb_excel.Close(False)
        excel.Quit()

        if os.path.exists(SOURCE_FILE_XLSX):
            return SOURCE_FILE_XLSX
    except Exception as excel_error:
        excel_conversion_error = excel_error

    # Method 2: LibreOffice installed and available in PATH
    libreoffice_commands = ["libreoffice", "soffice"]
    for command in libreoffice_commands:
        try:
            subprocess.run(
                [command, "--headless", "--convert-to", "xlsx", SOURCE_FILE_XLS, "--outdir", "."],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            if os.path.exists(SOURCE_FILE_XLSX):
                return SOURCE_FILE_XLSX
        except FileNotFoundError:
            continue
        except Exception as libreoffice_error:
            last_libreoffice_error = libreoffice_error

    raise RuntimeError(
        "Could not convert the SEBI .xls file to .xlsx.\n\n"
        "Fix option 1: Open 'SEBI_Monthly_Portfolio 31 JAN 2026.xls' in Excel, "
        "then Save As 'SEBI_Monthly_Portfolio 31 JAN 2026.xlsx', and run this script again.\n"
        "Fix option 2: Install pywin32 using: pip install pywin32, if Microsoft Excel is installed.\n"
        "Fix option 3: Install LibreOffice and add it to PATH.\n\n"
        f"Excel conversion error was: {excel_conversion_error}"
    )


def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\r", "\n").strip()


def normalize_isin(value):
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    return str(value).strip().upper()


def norm_header(value):
    """Normalize Excel headers so CRLF/LF/_x000D_/extra spaces do not break matching."""
    if value is None:
        return ""
    text = str(value)
    # When .xls is converted to .xlsx, Excel/openpyxl can expose carriage return as literal _x000D_.
    text = text.replace("_x000D_", "\n").replace("\r", "\n")
    lines = [part.strip() for part in text.split("\n") if part.strip()]
    return "\n".join(lines).strip()


def canonical_header(value):
    """Return the standard column name used by this script."""
    h = norm_header(value)
    flat = " ".join(h.replace("\n", " ").split()).casefold()

    aliases = {
        "name of the instrument": "Name of the Instrument",
        "isin": "ISIN",
        "rating": "Rating",
        "industry^ / rating": "Industry^ / Rating",
        "industry^/ rating": "Industry^ / Rating",
        "industry^/rating": "Industry^ / Rating",
        "quantity": "Quantity",
        "market/fair value (rs.in lacs)": "Market/Fair Value\n(Rs.in Lacs)",
        "market/fair value (rs. in lacs)": "Market/Fair Value\n(Rs.in Lacs)",
        "% to net assets": "% to Net Assets",
    }
    return aliases.get(flat, h)


def to_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, str):
        value = value.replace(",", "").replace("%", "").replace("$", "").strip()
        if value in {"", "-"}:
            return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    try:
        return float(value)
    except Exception:
        return None


def to_decimal(value):
    if value is None or value == "":
        return None

    if isinstance(value, str):
        value = value.replace(",", "").replace("%", "").replace("$", "").strip()
        if value in {"", "-"}:
            return None

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def round_half_up(value, decimals=2):
    dec_value = to_decimal(value)

    if dec_value is None:
        return None
    
    quant = Decimal("1").scaleb(-decimals)     #decimal=2 -> Decimal("0.01")
    return dec_value.quantize(quant, rounding="ROUND_HALF_UP")


def percent_value_for_compare(value, source):
    dec_value = to_decimal(value)

    if dec_value is None:
        return None

    if source == "SEBI":
        # SEBI percentage cells are stored internally as decimal values.
        # Example formula bar 6.46% is read by Python as 0.0646.
        return dec_value * Decimal("100")

    # Master already stores percentage as normal number.
    # Example 6.46 means 6.46%.
    return dec_value


def format_percent_for_index(value, source):
    if value is None:
        return ""
    
    raw_text = str(value).strip()

    # If SEBI already has a visible percentage string like "$0.00%"
    # show it exactly as it appears in the SEBI sheet.
    if source == "SEBI" and "%" in raw_text:
        return raw_text
    
    dec_value = to_decimal(value)

    if dec_value is None:
        return ""

    if source == "SEBI":
        dec_value = dec_value * Decimal("100")

    return format(dec_value, "f").rstrip("0").rstrip(".") + "%"


def round_to_2(dec_value):
    if dec_value is None:
        return None
    return dec_value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def same_value(master_value, sebi_value, column_name):
    # Quantity must be exact numeric comparison: no rounding, no conversion.
    if column_name == "Quantity":
        mv = to_decimal(master_value)
        sv = to_decimal(sebi_value)

        if mv is None and sv is None:
            return True
        if mv is None or sv is None:
            return False
        return mv == sv

    # % to Net Assets:
    # Master stores 6.46 as 6.46%, while SEBI stores 6.46% internally as 0.0646.
    # Convert SEBI * 100, then compare after ROUND_HALF_UP to 2 decimals.
    if column_name == "% to Net Assets":
        mv = percent_value_for_compare(master_value, "MASTER")
        sv = percent_value_for_compare(sebi_value, "SEBI")

        if mv is None and sv is None:
            return True
        if mv is None or sv is None:
            return False
        return round_to_2(mv) == round_to_2(sv)

    # Market/Fair Value and Yield compare after ROUND_HALF_UP to 2 decimals.
    # Text columns fall through to cleaned case-insensitive comparison.
    mv = to_decimal(master_value)
    sv = to_decimal(sebi_value)

    if mv is not None or sv is not None:
        if mv is None or sv is None:
            return False
        return round_to_2(mv) == round_to_2(sv)

    return clean_text(master_value).casefold() == clean_text(sebi_value).casefold()


def get_header_map(ws):
    """Find the SEBI header row and map columns by header names."""
    for row in range(1, ws.max_row + 1):
        values = [canonical_header(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if "Name of the Instrument" in values and "ISIN" in values:
            header_row = row
            col_map = {}
            for col in range(1, ws.max_column + 1):
                header = canonical_header(ws.cell(row, col).value)
                if header in SEBI_COLUMNS:
                    col_map[header] = col
            missing = [c for c in SEBI_COLUMNS if c not in col_map]
            if missing:
                available = [norm_header(ws.cell(row, col).value) for col in range(1, ws.max_column + 1) if ws.cell(row, col).value]
                raise ValueError(f"Missing required SEBI columns {missing}. Available headers: {available}")
            return header_row, col_map
    raise ValueError(f"Header row not found in sheet {ws.title}")


def find_first_equity_section_rows(ws, header_row, name_col):
    """
    For equity schemes only:
    Start after first '(a) Listed / awaiting listing on Stock Exchange'.
    Stop at the first 'Sub Total'.
    Also capture the first 'Total' after that Sub Total.
    Nothing below this first Total is checked.
    """
    start_row = None
    subtotal_row = None
    total_row = None

    for row in range(header_row + 1, ws.max_row + 1):
        label = clean_text(ws.cell(row, name_col).value)
        label_lower = label.casefold()
        if start_row is None and "(a) listed / awaiting listing on stock" in label_lower:
            start_row = row + 1
            continue
        if start_row is not None and subtotal_row is None and label_lower == "sub total":
            subtotal_row = row
            continue
        if subtotal_row is not None and label_lower == "total":
            total_row = row
            break

    if start_row is None or subtotal_row is None:
        raise ValueError(f"First equity Sub Total section not found in sheet {ws.title}")

    # If Total is missing, still compare Sub Total and only read rows before Sub Total.
    return start_row, subtotal_row, total_row


def read_sebi_first_section(ws):
    header_row, col_map = get_header_map(ws)
    name_col = col_map["Name of the Instrument"]
    start_row, subtotal_row, total_row = find_first_equity_section_rows(ws, header_row, name_col)

    records = []
    row_by_isin = {}
    for row in range(start_row, subtotal_row):
        isin = normalize_isin(ws.cell(row, col_map["ISIN"]).value)
        name = clean_text(ws.cell(row, name_col).value)
        if not isin and not name:
            continue
        rec = {"Excel Row": row}
        for col in SEBI_COLUMNS:
            rec[col] = ws.cell(row, col_map[col]).value
        records.append(rec)
        if isin:
            row_by_isin[isin] = rec

    subtotal = {"Excel Row": subtotal_row}
    for col in SEBI_COLUMNS:
        subtotal[col] = ws.cell(subtotal_row, col_map[col]).value

    total = None
    if total_row:
        total = {"Excel Row": total_row}
        for col in SEBI_COLUMNS:
            total[col] = ws.cell(total_row, col_map[col]).value

    return {
        "header_row": header_row,
        "col_map": col_map,
        "records": records,
        "row_by_isin": row_by_isin,
        "subtotal": subtotal,
        "total": total,
    }


def read_master_equity(master_df, scheme_code):
    master_table = master_df[
        (master_df["Client Code"] == scheme_code)
        & (master_df["Security Type Name"].isin(["EQUITY", "PREFERRED STOCK"]))
    ].copy()

    master_table = master_table[
        [
            "Client Code",
            "Issuer Name",
            "Security Type Name",
            "ISIN",
            "Industry",
            "Quantity",
            "Total Market Value (Rs.)",
            "% to Net assests",
        ]
    ]

    master_table = master_table.rename(columns=MASTER_RENAME)
    master_table["Market/Fair Value\n(Rs.in Lacs)"] = master_table[
        "Market/Fair Value\n(Rs.in Lacs)"
    ].apply(
        lambda x: round_half_up(
            Decimal(str(x)) / Decimal(str(MARKET_VALUE_DIVISOR)),
            2
        ) if pd.notna(x) else None
    )

    # Master stores % as 1.23 for 1.23%.
    # Do NOT divide by 100.
    # SEBI percentage is converted only during comparison/output.
    master_table["% to Net Assets"] = pd.to_numeric(
        master_table["% to Net Assets"], errors="coerce"
    )

    master_table = master_table[
        [
            "Name of the Instrument",
            "ISIN",
            "Industry^ / Rating",
            "Quantity",
            "Market/Fair Value\n(Rs.in Lacs)",
            "% to Net Assets",
        ]
    ].reset_index(drop=True)

    records = master_table.to_dict("records")
    by_isin = {normalize_isin(r["ISIN"]): r for r in records if normalize_isin(r.get("ISIN"))}
    return master_table, records, by_isin


def add_discrepancy(discrepancies, scheme, row_type, isin, instrument, column, master_value, sebi_value, status, excel_row=None):
    discrepancies.append(
        {
            "Scheme Code": scheme,
            "Row Type": row_type,
            "ISIN": isin,
            "Name of the Instrument": instrument,
            "Column": column,
            "Master Value": master_value,
            "SEBI Value": sebi_value,
            "Status": status,
            "SEBI Excel Row": excel_row,
        }
    )


def highlight_cell(ws, row, col, fill):
    if row and col:
        ws.cell(row, col).fill = fill


def compare_scheme(ws, scheme_code, master_df, all_discrepancies):
    sebi = read_sebi_first_section(ws)
    master_table, master_records, master_by_isin = read_master_equity(master_df, scheme_code)
    sebi_by_isin = sebi["row_by_isin"]
    col_map = sebi["col_map"]

    # 1. Security row comparison by ISIN.
    all_isins = sorted(set(master_by_isin.keys()) | set(sebi_by_isin.keys()))
    for isin in all_isins:
        master_row = master_by_isin.get(isin)
        sebi_row = sebi_by_isin.get(isin)

        if master_row is None:
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                "Security Row",
                isin,
                sebi_row.get("Name of the Instrument"),
                "Entire Row",
                None,
                "Present",
                "Master workbook does not have this entire row",
                sebi_row.get("Excel Row"),
            )
            continue

        if sebi_row is None:
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                "Security Row",
                isin,
                master_row.get("Name of the Instrument"),
                "Entire Row",
                "Present",
                None,
                "SEBI sheet does not have this entire row",
                None,
            )
            continue

        for col in COMPARE_COLS:
            if not same_value(master_row.get(col), sebi_row.get(col), col):
                add_discrepancy(
                    all_discrepancies,
                    scheme_code,
                    "Security Row",
                    isin,
                    master_row.get("Name of the Instrument") or sebi_row.get("Name of the Instrument"),
                    col,
                    master_row.get(col),
                    sebi_row.get(col),
                    "Value mismatch",
                    sebi_row.get("Excel Row"),
                )
                highlight_cell(ws, sebi_row.get("Excel Row"), col_map.get(col), HIGHLIGHT_FILL)

    # 2. First Sub Total and first Total comparison.
    master_market_total = round(pd.to_numeric(master_table["Market/Fair Value\n(Rs.in Lacs)"], errors="coerce").fillna(0).sum(), 2)
    master_percent_total = pd.to_numeric(master_table["% to Net Assets"], errors="coerce").fillna(0).sum()

    for row_label, sebi_total_row in [("Sub Total", sebi["subtotal"]), ("Total", sebi["total"])] :
        if not sebi_total_row:
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                row_label,
                "",
                row_label,
                "Entire Row",
                "Present",
                None,
                f"SEBI sheet does not have first {row_label} row",
                None,
            )
            continue

        row_no = sebi_total_row.get("Excel Row")
        market_col = "Market/Fair Value\n(Rs.in Lacs)"
        pct_col = "% to Net Assets"

        if not same_value(master_market_total, sebi_total_row.get(market_col), market_col):
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                row_label,
                "",
                row_label,
                market_col,
                master_market_total,
                sebi_total_row.get(market_col),
                "Value mismatch",
                row_no,
            )
            highlight_cell(ws, row_no, col_map.get(market_col), TOTAL_FILL)

        if not same_value(master_percent_total, sebi_total_row.get(pct_col), pct_col):
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                row_label,
                "",
                row_label,
                pct_col,
                master_percent_total,
                sebi_total_row.get(pct_col),
                "Value mismatch",
                row_no,
            )
            highlight_cell(ws, row_no, col_map.get(pct_col), TOTAL_FILL)


def get_header_map_for_columns(ws, required_columns):
    """Find a header row for the provided column list and return a column map."""
    for row in range(1, ws.max_row + 1):
        values = [canonical_header(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]

        # Debt files sometimes use "Industry^ / Rating" in the rating position.
        debt_values = ["Rating" if value == "Industry^ / Rating" else value for value in values]

        if "Name of the Instrument" in debt_values and "ISIN" in debt_values:
            header_row = row
            col_map = {}
            for col in range(1, ws.max_column + 1):
                header = canonical_header(ws.cell(row, col).value)
                if "Rating" in required_columns and header == "Industry^ / Rating":
                    header = "Rating"
                if header in required_columns:
                    col_map[header] = col

            missing = [c for c in required_columns if c not in col_map]
            if missing:
                available = [norm_header(ws.cell(row, col).value) for col in range(1, ws.max_column + 1) if ws.cell(row, col).value]
                raise ValueError(f"Missing required SEBI columns {missing}. Available headers: {available}")
            return header_row, col_map

    raise ValueError(f"Header row not found in sheet {ws.title}")


def read_sebi_debt_sections(ws):
    """
    Read all debt security rows from the SEBI sheet from header until GRAND TOTAL.
    Security rows are identified by non-empty normalized ISIN.

    It also captures every Sub Total and Total row, with the ISINs that appeared
    in that subtotal/total block, so totals can be checked without hard-coding
    every debt subsection name.
    """
    header_row, col_map = get_header_map_for_columns(ws, DEBT_SEBI_COLUMNS)
    name_col = col_map["Name of the Instrument"]

    records = []
    row_by_isin = {}
    subtotal_rows = []
    total_rows = []

    current_subtotal_isins = []
    current_total_isins = []
    all_isins_before_grand_total = []

    for row in range(header_row + 1, ws.max_row + 1):
        label = clean_text(ws.cell(row, name_col).value)
        label_lower = label.casefold()

        if label_lower == "grand total":
            grand_total = {"Excel Row": row, "ISINs": list(all_isins_before_grand_total)}
            for col in DEBT_SEBI_COLUMNS:
                grand_total[col] = ws.cell(row, col_map[col]).value
            return {
                "header_row": header_row,
                "col_map": col_map,
                "records": records,
                "row_by_isin": row_by_isin,
                "subtotals": subtotal_rows,
                "totals": total_rows,
                "grand_total": grand_total,
            }

        if label_lower == "sub total":
            subtotal = {"Excel Row": row, "ISINs": list(current_subtotal_isins)}
            for col in DEBT_SEBI_COLUMNS:
                subtotal[col] = ws.cell(row, col_map[col]).value
            subtotal_rows.append(subtotal)
            current_subtotal_isins = []
            continue

        if label_lower == "total":
            total = {"Excel Row": row, "ISINs": list(current_total_isins)}
            for col in DEBT_SEBI_COLUMNS:
                total[col] = ws.cell(row, col_map[col]).value
            total_rows.append(total)
            current_total_isins = []
            continue

        isin = normalize_isin(ws.cell(row, col_map["ISIN"]).value)
        name = clean_text(ws.cell(row, name_col).value)

        # Section headings/subsection headings have no ISIN; do not compare them as securities.
        if not isin:
            continue

        rec = {"Excel Row": row}
        for col in DEBT_SEBI_COLUMNS:
            rec[col] = ws.cell(row, col_map[col]).value
        records.append(rec)
        row_by_isin[isin] = rec
        current_subtotal_isins.append(isin)
        current_total_isins.append(isin)
        all_isins_before_grand_total.append(isin)

    raise ValueError(f"GRAND TOTAL row not found in debt sheet {ws.title}")


def read_master_debt(master_df, scheme_code):
    master_table = master_df[
        (master_df["Client Code"] == scheme_code)
        & (master_df["Security Type Name"].isin(DEBT_MASTER_SECURITY_TYPES))
    ].copy()

    required_master_cols = [
        "Client Code",
        "Security Name",
        "Security Type Name",
        "ISIN",
        "Rating",
        "Quantity",
        "Total Market Value (Rs.)",
        "% to Net assests",
        "Yield",
    ]
    missing = [c for c in required_master_cols if c not in master_table.columns]
    if missing:
        raise ValueError(f"Missing required master columns for debt {missing}")

    master_table = master_table[required_master_cols]
    master_table = master_table.rename(columns=DEBT_MASTER_RENAME)

    # Debt Market/Fair Value follows same rule as equity:
    # Master Total Market Value (Rs.) / 100000, then ROUND_HALF_UP to 2 decimals.
    master_table["Market/Fair Value\n(Rs.in Lacs)"] = master_table[
        "Market/Fair Value\n(Rs.in Lacs)"
    ].apply(
        lambda x: round_half_up(
            Decimal(str(x)) / Decimal(str(MARKET_VALUE_DIVISOR)),
            2
        ) if pd.notna(x) else None
    )

    # Master stores % as percent number, e.g. 6.46 means 6.46%.
    # Do not divide by 100; SEBI is converted only during comparison/output.
    master_table["% to Net Assets"] = pd.to_numeric(
        master_table["% to Net Assets"], errors="coerce"
    )

    master_table = master_table[
        [
            "Name of the Instrument",
            "ISIN",
            "Rating",
            "Quantity",
            "Market/Fair Value\n(Rs.in Lacs)",
            "% to Net Assets",
            "Yield",
        ]
    ].reset_index(drop=True)

    records = master_table.to_dict("records")
    by_isin = {normalize_isin(r["ISIN"]): r for r in records if normalize_isin(r.get("ISIN"))}
    return master_table, records, by_isin


def sum_master_column_by_isins(master_by_isin, isins, column_name):
    total = Decimal("0")
    found_any = False

    for isin in isins:
        row = master_by_isin.get(isin)
        if not row:
            continue

        value = row.get(column_name)
        if column_name == "% to Net Assets":
            dec_value = percent_value_for_compare(value, "MASTER")
        else:
            dec_value = to_decimal(value)

        if dec_value is not None:
            total += dec_value
            found_any = True

    if not found_any:
        return None

    if column_name == "Market/Fair Value\n(Rs.in Lacs)":
        return round_to_2(total)

    return total


def compare_debt_total_row(ws, scheme_code, row_type, total_row, master_by_isin, col_map, all_discrepancies):
    row_no = total_row.get("Excel Row")
    market_col = "Market/Fair Value\n(Rs.in Lacs)"
    pct_col = "% to Net Assets"

    master_market_total = sum_master_column_by_isins(master_by_isin, total_row.get("ISINs", []), market_col)
    master_percent_total = sum_master_column_by_isins(master_by_isin, total_row.get("ISINs", []), pct_col)

    if not same_value(master_market_total, total_row.get(market_col), market_col):
        add_discrepancy(
            all_discrepancies,
            scheme_code,
            row_type,
            "",
            row_type,
            market_col,
            master_market_total,
            total_row.get(market_col),
            "Value mismatch",
            row_no,
        )
        highlight_cell(ws, row_no, col_map.get(market_col), TOTAL_FILL)

    if not same_value(master_percent_total, total_row.get(pct_col), pct_col):
        add_discrepancy(
            all_discrepancies,
            scheme_code,
            row_type,
            "",
            row_type,
            pct_col,
            master_percent_total,
            total_row.get(pct_col),
            "Value mismatch",
            row_no,
        )
        highlight_cell(ws, row_no, col_map.get(pct_col), TOTAL_FILL)


def compare_scheme_debt(ws, scheme_code, master_df, all_discrepancies):
    sebi = read_sebi_debt_sections(ws)
    master_table, master_records, master_by_isin = read_master_debt(master_df, scheme_code)
    sebi_by_isin = sebi["row_by_isin"]
    col_map = sebi["col_map"]

    # 1. Security row comparison by normalized ISIN.
    # Missing whole rows are reported in Index only; no highlighting for missing rows.
    all_isins = sorted(set(master_by_isin.keys()) | set(sebi_by_isin.keys()))
    for isin in all_isins:
        master_row = master_by_isin.get(isin)
        sebi_row = sebi_by_isin.get(isin)

        if master_row is None:
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                "Security Row",
                isin,
                sebi_row.get("Name of the Instrument"),
                "Entire Row",
                None,
                "Present",
                "Master workbook does not have this entire row",
                sebi_row.get("Excel Row"),
            )
            continue

        if sebi_row is None:
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                "Security Row",
                isin,
                master_row.get("Name of the Instrument"),
                "Entire Row",
                "Present",
                None,
                "SEBI sheet does not have this entire row",
                None,
            )
            continue

        for col in DEBT_COMPARE_COLS:
            if not same_value(master_row.get(col), sebi_row.get(col), col):
                add_discrepancy(
                    all_discrepancies,
                    scheme_code,
                    "Security Row",
                    isin,
                    master_row.get("Name of the Instrument") or sebi_row.get("Name of the Instrument"),
                    col,
                    master_row.get(col),
                    sebi_row.get(col),
                    "Value mismatch",
                    sebi_row.get("Excel Row"),
                )
                highlight_cell(ws, sebi_row.get("Excel Row"), col_map.get(col), HIGHLIGHT_FILL)

    # 2. Debt Sub Total and Total rows.
    # Each row is checked against the master sum for the ISINs present in that SEBI block.
    for idx, subtotal_row in enumerate(sebi["subtotals"], start=1):
        compare_debt_total_row(
            ws,
            scheme_code,
            f"Sub Total {idx}",
            subtotal_row,
            master_by_isin,
            col_map,
            all_discrepancies,
        )

    for idx, total_row in enumerate(sebi["totals"], start=1):
        compare_debt_total_row(
            ws,
            scheme_code,
            f"Total {idx}",
            total_row,
            master_by_isin,
            col_map,
            all_discrepancies,
        )


def write_index_output(wb, discrepancies):
    if "Index" not in wb.sheetnames:
        ws = wb.create_sheet("Index", 0)
    else:
        ws = wb["Index"]

    # Clear old output from E onwards only.
    for row in range(1, ws.max_row + 1):
        for col in range(5, max(ws.max_column, 20) + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.font = copy(ws.cell(row, 1).font)
            cell.border = Border()
            cell.alignment = Alignment()

    title_cell = ws.cell(1, 5)
    title_cell.value = "Equity + Debt Reconciliation Output"
    title_cell.font = Font(bold=True, size=12)

    headers = [
        "Scheme Code",
        "Row Type",
        "ISIN",
        "Name of the Instrument",
        "Column",
        "Master Value",
        "SEBI Value",
        "Status",
        "SEBI Excel Row",
    ]

    start_row = 3
    start_col = 5
    for i, header in enumerate(headers, start=start_col):
        cell = ws.cell(start_row, i)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for r_idx, rec in enumerate(discrepancies, start=start_row + 1):
        for c_idx, header in enumerate(headers, start=start_col):
            cell = ws.cell(r_idx, c_idx)

            if rec.get("Column") == "% to Net Assets" and header == "Master Value":
                cell.value = format_percent_for_index(rec.get("Master Value"), "MASTER")

            elif rec.get("Column") == "% to Net Assets" and header == "SEBI Value":
                cell.value = format_percent_for_index(rec.get("SEBI Value"), "SEBI")

            else:
                cell.value = rec.get(header)

            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        5: 14, 6: 16, 7: 18, 8: 42, 9: 30, 10: 20, 11: 20, 12: 42, 13: 14
    }
    for col, width in widths.items():
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    ws.freeze_panes = "E4"
    ws.auto_filter.ref = f"E3:M{max(start_row + 1, start_row + len(discrepancies))}"


def main():
    source_xlsx = resolve_source_xlsx()

    wb = load_workbook(source_xlsx)
    master_df = pd.read_excel(MASTER_FILE)

    discrepancies = []
    processed = []
    skipped = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Index":
            continue

        scheme_type = scheme_dict.get(sheet_name)
        if scheme_type not in {"EQUITY", "DEBT"}:
            continue

        ws = wb[sheet_name]
        try:
            if scheme_type == "EQUITY":
                compare_scheme(ws, sheet_name, master_df, discrepancies)
            elif scheme_type == "DEBT":
                compare_scheme_debt(ws, sheet_name, master_df, discrepancies)
            processed.append(f"{sheet_name} ({scheme_type})")
        except Exception as exc:
            skipped.append((sheet_name, str(exc)))
            add_discrepancy(
                discrepancies,
                sheet_name,
                "Sheet Error",
                "",
                "",
                "",
                "",
                "",
                f"Could not process sheet: {exc}",
                None,
            )

    write_index_output(wb, discrepancies)
    wb.save(OUTPUT_FILE)

    print(f"Processed equity/debt schemes: {len(processed)}")
    print(f"Discrepancies written: {len(discrepancies)}")
    if skipped:
        print("Skipped sheets:")
        for item in skipped:
            print(item)
    print(f"Output file: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()