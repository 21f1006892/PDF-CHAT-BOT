import os
import math
import subprocess
from copy import copy
from collections import defaultdict
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


# ===================== FILE PATHS =====================
SOURCE_FILE_XLS = "SEBI_Monthly_Portfolio 31 JAN 2026.xls"
SOURCE_FILE_XLSX = "SEBI_Monthly_Portfolio 31 JAN 2026.xlsx"
MASTER_FILE = "IN_MF_PORTFOLIO_DETAILS_ACROSS_SCHEMES.xlsx"
TRIAL_BALANCE_FILE = "CITI_ABC_Trial_Balance_310126.xlsx"
TRIAL_BALANCE_SHEET = "sg_in003_Y4X_300126_310126"
OUTPUT_FILE = "SEBI_Monthly_Portfolio_Equity_Remaining_Sections_Reconciled.xlsx"

# IMPORTANT:
# In the uploaded master file, Total Market Value (Rs.) matches SEBI's
# "Market/Fair Value\n(Rs.in Lacs)" after dividing by 100000 because 1 lac = 100000.
MARKET_VALUE_DIVISOR = 100000

# ===================== SCHEME TYPE DICTIONARY FROM TESTTWO.PY =====================
scheme_dict = {'ABBSEIIF': 'EQUITY', 'ABSLBCF': 'EQUITY', 'ABSLCONF': 'EQUITY', 'ABSLESG': 'EQUITY',
'ABSLLDF': 'DEBT', 'ABSLMAAF': 'HYBRID', 'ABSLMCF': 'EQUITY', 'ABSLQF': 'EQUITY', 'ABSLSO': 'EQUITY',
'ABSLTNLF': 'EQUITY', 'ABSLUS03': 'FOF', 'ABSLUS10': 'FOF', 'ADVG': 'EQUITY', 'BANKETF': 'EQUITY', 'BBIF': 'DEBT',
'BBP': 'DEBT', 'BDB': 'DEBT', 'BDYP': 'EQUITY', 'BFL': 'DEBT', 'BFS': 'DEBT', 'BINFRA': 'EQUITY',
'BINTEQA': 'INTRNL', 'BMIDX50': 'EQUITY', 'BQIDX50': 'EQUITY', 'BSL95F': 'HYBRID', 'BSLAAMM': 'FOF',
'BSLADMM': 'FOF', 'BSLBBYW': 'EQUITY', 'BSLBKFS': 'EQUITY', 'BSLCBF': 'DEBT', 'BSLCM': 'DEBT', 'BSLDAAF': 'HYBRID',
'BSLEAF': 'HYBRID', 'BSLEQSF': 'HYBRID', 'BSLEQTY': 'EQUITY', 'BSLFEF': 'HYBRID', 'BSLFPAP': 'FOF', 'BSLFPCP': 'FOF',
'BSLFPPP': 'FOF', 'BSLIF': 'DEBT', 'BSLMFG': 'EQUITY', 'BSLMIFOF': 'FOF', 'BSLMTP': 'DEBT', 'BSLNMF': 'EQUITY',
'BSLONF': 'DEBT', 'BSLPHF': 'EQUITY', 'BSLR96': 'EQUITY', 'BSLRF30': 'HYBRID', 'BSLRF40': 'HYBRID',
'BSLRF50': 'HYBRID', 'BSLRF50P': 'DEBT', 'BSLSTF': 'DEBT', 'BSLTA1': 'EQUITY', 'BTOP100': 'EQUITY',
'C10YGETF': 'DEBT', 'CASH': 'DEBT', 'CBGETF': 'DEBT', 'CIGAPR26': 'DEBT', 'CIGAPR28': 'DEBT', 'CIGAPR29': 'DEBT',
'CIGAPR33': 'DEBT', 'CIGJUN27': 'DEBT', 'CIGPSA28': 'DEBT', 'CISJUN32': 'DEBT', 'COFAIETF': 'DEBT',
'CSFSD12': 'DEBT', 'CSFSD6': 'DEBT', 'CSFSI27': 'DEBT', 'CSNHFS26': 'DEBT', 'CSPAPA26': 'DEBT', 'CSPAPA27': 'DEBT',
'FTPTI': 'DEBT', 'FTPTJ': 'DEBT', 'FTPTQ': 'DEBT', 'FTPUB': 'DEBT', 'FTPUJ': 'DEBT', 'GENNEXT': 'EQUITY',
'GOLDETF': 'GOLD', 'GOLDFOF': 'OTH', 'INV': 'DEBT', 'MIDCAP': 'EQUITY', 'MIP25': 'HYBRID', 'MNC': 'EQUITY',
'N30MTETF': 'EQUITY', 'N30QTETF': 'EQUITY', 'NEWIF50': 'EQUITY', 'NHLTHETF': 'EQUITY', 'NIFTY': 'EQUITY',
'NIFTYETF': 'EQUITY', 'NIFYNX50': 'EQUITY', 'NINDDEF': 'EQUITY', 'NITETF': 'EQUITY', 'NMID150': 'EQUITY',
'NPSEETF': 'EQUITY', 'NSDAPR27': 'DEBT', 'NSDAQ100': 'FOF', 'NSDSEP27': 'DEBT', 'NSMALL50': 'EQUITY',
'NSPPBS26': 'DEBT', 'NXTIDX50': 'EQUITY', 'PLUS': 'DEBT', 'PSUEQ': 'EQUITY', 'PURE': 'EQUITY', 'SENSXETF': 'EQUITY',
'SILVRETF': 'SILVER', 'SILVRFOF': 'OTH', 'BSLGCF': 'FOF', 'BSLGRE': 'FOF'}

SEBI_COLUMNS = [
    "Name of the Instrument",
    "ISIN",
    "Industry^ / Rating",
    "Quantity",
    "Market/Fair Value\n(Rs.in Lacs)",
    "% to Net Assets",
]

MASTER_RENAME = {
    "Issuer Name": "Name of the Instrument",
    "Industry": "Industry^ / Rating",
    "Total Market Value (Rs.)": "Market/Fair Value\n(Rs.in Lacs)",
    "% to Net assests": "% to Net Assets",
}

COMPARE_COLS = [
    "Name of the Instrument",
    "Industry^ / Rating",
    "Quantity",
    "Market/Fair Value\n(Rs.in Lacs)",
    "% to Net Assets",
]


# Remaining sections to be checked for EQUITY schemes after the first Total.
# "Others" is only a parent section; actual checking happens for child rows/sections under it.
REMAINING_EQUITY_SECTIONS = {
    "Others",
    "TREPS / Reverse Repo",
    "Foreign Securities and/or overseas ETF(s)",
    "(a) Listed / awaiting listing on Stock Exchanges",
    "International Mutual Fund Units",
    "Exchange Traded Funds",
    "Equity Futures/Index",
    "Margin (Future and Options)",
    "Cash and Bank",
}

# These two sections are checked from Trial Balance, not from the main Master workbook.
TRIAL_BALANCE_ACCOUNT_CODES = {
    "Cash and Bank": "141839",
    "Margin (Future and Options)": "141350",
}

# Master Security Type Name mapping for remaining sections.
# The strings are normalized before comparison, so extra spaces/case differences are handled.
REMAINING_SECTION_SECURITY_TYPES = {
    "Foreign Securities and/or overseas ETF(s)": ["International Equity"],
    "International Mutual Fund Units": ["INVESTMENT FUNDS/MUTUAL FUNDS"],
    "Exchange Traded Funds": ["Exchange Traded Fund"],
    "TREPS / Reverse Repo": ["COLLATERALISED BORROWING AND LENDING", "COLLATERALISED BORROWING AND LENDIN", "REPO"],
    "Equity Futures/Index": ["EQUITY  FUTURE", "EQUITY FUTURE"],
}

REMAINING_COMPARE_COLS = [
    "Name of the Instrument",
    "Industry^ / Rating",
    "Quantity",
    "Market/Fair Value\n(Rs.in Lacs)",
    "% to Net Assets",
]


HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")     # yellow
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")         # light red
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def resolve_source_xlsx():
    """
    Return an xlsx file path that openpyxl can edit.

    openpyxl cannot edit old .xls files directly.
    So this function first looks for an already-saved .xlsx file.
    If only .xls exists, it tries:
      1) Microsoft Excel COM conversion on Windows, if Excel + pywin32 are available.
      2) LibreOffice conversion, if LibreOffice is installed and available in PATH.

    Easiest manual fix if conversion fails:
      Open the SEBI .xls file in Excel -> File -> Save As -> .xlsx
      and keep the file name as SOURCE_FILE_XLSX below.
    """
    if os.path.exists(SOURCE_FILE_XLSX):
        return SOURCE_FILE_XLSX

    if not os.path.exists(SOURCE_FILE_XLS):
        raise FileNotFoundError(
            f"Could not find {SOURCE_FILE_XLSX} or {SOURCE_FILE_XLS} in this folder."
        )

    # Method 1: Windows + Microsoft Excel installed
    try:
        import win32com.client  # pip install pywin32

        src = os.path.abspath(SOURCE_FILE_XLS)
        dst = os.path.abspath(SOURCE_FILE_XLSX)

        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_excel = excel.Workbooks.Open(src)
        # 51 = xlOpenXMLWorkbook = .xlsx
        wb_excel.SaveAs(dst, FileFormat=51)
        wb_excel.Close(False)
        excel.Quit()

        if os.path.exists(SOURCE_FILE_XLSX):
            return SOURCE_FILE_XLSX
    except Exception as excel_error:
        excel_conversion_error = excel_error

    # Method 2: LibreOffice installed and available in PATH
    libreoffice_commands = ["libreoffice", "soffice"]
    for command in libreoffice_commands:
        try:
            subprocess.run(
                [command, "--headless", "--convert-to", "xlsx", SOURCE_FILE_XLS, "--outdir", "."],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            if os.path.exists(SOURCE_FILE_XLSX):
                return SOURCE_FILE_XLSX
        except FileNotFoundError:
            continue
        except Exception as libreoffice_error:
            last_libreoffice_error = libreoffice_error

    raise RuntimeError(
        "Could not convert the SEBI .xls file to .xlsx.\n\n"
        "Fix option 1: Open 'SEBI_Monthly_Portfolio 31 JAN 2026.xls' in Excel, "
        "then Save As 'SEBI_Monthly_Portfolio 31 JAN 2026.xlsx', and run this script again.\n"
        "Fix option 2: Install pywin32 using: pip install pywin32, if Microsoft Excel is installed.\n"
        "Fix option 3: Install LibreOffice and add it to PATH.\n\n"
        f"Excel conversion error was: {excel_conversion_error}"
    )


def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\r", "\n").strip()


def normalize_isin(value):
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    return str(value).strip().upper()


def norm_header(value):
    """Normalize Excel headers so CRLF/LF/_x000D_/extra spaces do not break matching."""
    if value is None:
        return ""
    text = str(value)
    # When .xls is converted to .xlsx, Excel/openpyxl can expose carriage return as literal _x000D_.
    text = text.replace("_x000D_", "\n").replace("\r", "\n")
    lines = [part.strip() for part in text.split("\n") if part.strip()]
    return "\n".join(lines).strip()


def canonical_header(value):
    """Return the standard column name used by this script."""
    h = norm_header(value)
    flat = " ".join(h.replace("\n", " ").split()).casefold()

    aliases = {
        "name of the instrument": "Name of the Instrument",
        "isin": "ISIN",
        "industry^ / rating": "Industry^ / Rating",
        "industry^/ rating": "Industry^ / Rating",
        "industry^/rating": "Industry^ / Rating",
        "quantity": "Quantity",
        "market/fair value (rs.in lacs)": "Market/Fair Value\n(Rs.in Lacs)",
        "market/fair value (rs. in lacs)": "Market/Fair Value\n(Rs.in Lacs)",
        "% to net assets": "% to Net Assets",
    }
    return aliases.get(flat, h)


def to_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, str):
        value = value.replace(",", "").replace("%", "").replace("$", "").strip()
        if value in {"", "-"}:
            return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    try:
        return float(value)
    except Exception:
        return None


def to_decimal(value):
    if value is None or value == "":
        return None

    if isinstance(value, str):
        value = value.replace(",", "").replace("%", "").replace("$", "").strip()
        if value in {"", "-"}:
            return None

    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def round_half_up(value, decimals=2):
    dec_value = to_decimal(value)

    if dec_value is None:
        return None
    
    quant = Decimal("1").scaleb(-decimals)     #decimal=2 -> Decimal("0.01")
    return dec_value.quantize(quant, rounding="ROUND_HALF_UP")


def percent_value_for_compare(value, source):
    dec_value = to_decimal(value)

    if dec_value is None:
        return None

    if source == "SEBI":
        # SEBI percentage cells are stored internally as decimal values.
        # Example formula bar 6.46% is read by Python as 0.0646.
        return dec_value * Decimal("100")

    # Master already stores percentage as normal number.
    # Example 6.46 means 6.46%.
    return dec_value


def format_percent_for_index(value, source):
    if value is None:
        return ""
    
    raw_text = str(value).strip()

    # If SEBI already has a visible percentage string like "$0.00%"
    # show it exactly as it appears in the SEBI sheet.
    if source == "SEBI" and "%" in raw_text:
        return raw_text
    
    dec_value = to_decimal(value)

    if dec_value is None:
        return ""

    if source == "SEBI":
        dec_value = dec_value * Decimal("100")

    return format(dec_value, "f").rstrip("0").rstrip(".") + "%"


def round_to_2(dec_value):
    if dec_value is None:
        return None
    return dec_value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def same_value(master_value, sebi_value, column_name):
    # Compare numeric values after rounding to 2 decimal places.

    if column_name == "% to Net Assets":
        mv = percent_value_for_compare(master_value, "MASTER")
        sv = percent_value_for_compare(sebi_value, "SEBI")

        if mv is None and sv is None:
            return True

        if mv is None or sv is None:
            return False

        return round_to_2(mv) == round_to_2(sv)

    mv = to_decimal(master_value)
    sv = to_decimal(sebi_value)

    if mv is not None or sv is not None:
        if mv is None or sv is None:
            return False

        return round_to_2(mv) == round_to_2(sv)

    return clean_text(master_value).casefold() == clean_text(sebi_value).casefold()


def get_header_map(ws):
    """Find the SEBI header row and map columns by header names."""
    for row in range(1, ws.max_row + 1):
        values = [canonical_header(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if "Name of the Instrument" in values and "ISIN" in values:
            header_row = row
            col_map = {}
            for col in range(1, ws.max_column + 1):
                header = canonical_header(ws.cell(row, col).value)
                if header in SEBI_COLUMNS:
                    col_map[header] = col
            missing = [c for c in SEBI_COLUMNS if c not in col_map]
            if missing:
                available = [norm_header(ws.cell(row, col).value) for col in range(1, ws.max_column + 1) if ws.cell(row, col).value]
                raise ValueError(f"Missing required SEBI columns {missing}. Available headers: {available}")
            return header_row, col_map
    raise ValueError(f"Header row not found in sheet {ws.title}")


def find_first_equity_section_rows(ws, header_row, name_col):
    """
    For equity schemes only:
    Start after first '(a) Listed / awaiting listing on Stock Exchange'.
    Stop at the first 'Sub Total'.
    Also capture the first 'Total' after that Sub Total.
    Nothing below this first Total is checked.
    """
    start_row = None
    subtotal_row = None
    total_row = None

    for row in range(header_row + 1, ws.max_row + 1):
        label = clean_text(ws.cell(row, name_col).value)
        label_lower = label.casefold()
        if start_row is None and "(a) listed / awaiting listing on stock" in label_lower:
            start_row = row + 1
            continue
        if start_row is not None and subtotal_row is None and label_lower == "sub total":
            subtotal_row = row
            continue
        if subtotal_row is not None and label_lower == "total":
            total_row = row
            break

    if start_row is None or subtotal_row is None:
        raise ValueError(f"First equity Sub Total section not found in sheet {ws.title}")

    # If Total is missing, still compare Sub Total and only read rows before Sub Total.
    return start_row, subtotal_row, total_row


def read_sebi_first_section(ws):
    header_row, col_map = get_header_map(ws)
    name_col = col_map["Name of the Instrument"]
    start_row, subtotal_row, total_row = find_first_equity_section_rows(ws, header_row, name_col)

    records = []
    row_by_isin = {}
    for row in range(start_row, subtotal_row):
        isin = clean_text(ws.cell(row, col_map["ISIN"]).value)
        name = clean_text(ws.cell(row, name_col).value)
        if not isin and not name:
            continue
        rec = {"Excel Row": row}
        for col in SEBI_COLUMNS:
            rec[col] = ws.cell(row, col_map[col]).value
        records.append(rec)
        if isin:
            row_by_isin[isin] = rec

    subtotal = {"Excel Row": subtotal_row}
    for col in SEBI_COLUMNS:
        subtotal[col] = ws.cell(subtotal_row, col_map[col]).value

    total = None
    if total_row:
        total = {"Excel Row": total_row}
        for col in SEBI_COLUMNS:
            total[col] = ws.cell(total_row, col_map[col]).value

    return {
        "header_row": header_row,
        "col_map": col_map,
        "records": records,
        "row_by_isin": row_by_isin,
        "subtotal": subtotal,
        "total": total,
    }


def read_master_equity(master_df, scheme_code):
    master_table = master_df[
        (master_df["Client Code"] == scheme_code)
        & (master_df["Security Type Name"].isin(["EQUITY", "PREFERRED STOCK"]))
    ].copy()

    master_table = master_table[
        [
            "Client Code",
            "Issuer Name",
            "Security Type Name",
            "ISIN",
            "Industry",
            "Quantity",
            "Total Market Value (Rs.)",
            "% to Net assests",
        ]
    ]

    master_table = master_table.rename(columns=MASTER_RENAME)
    master_table["Market/Fair Value\n(Rs.in Lacs)"] = master_table[
        "Market/Fair Value\n(Rs.in Lacs)"
        ].apply(
            lambda x: round_to_2(Decimal(str(x)) / Decimal(str(MARKET_VALUE_DIVISOR))) 
            if pd.notna(x) else None
        )

    # Master stores % as 1.23 for 1.23%.
    # Do NOT divide by 100.
    # SEBI percentage is converted only during comparison/output.
    master_table["% to Net Assets"] = pd.to_numeric(
        master_table["% to Net Assets"], errors="coerce"
    )

    master_table = master_table[
        [
            "Name of the Instrument",
            "ISIN",
            "Industry^ / Rating",
            "Quantity",
            "Market/Fair Value\n(Rs.in Lacs)",
            "% to Net Assets",
        ]
    ].reset_index(drop=True)

    records = master_table.to_dict("records")
    by_isin = {clean_text(r["ISIN"]): r for r in records if clean_text(r.get("ISIN"))}
    return master_table, records, by_isin


def add_discrepancy(discrepancies, scheme, row_type, isin, instrument, column, master_value, sebi_value, status, excel_row=None):
    discrepancies.append(
        {
            "Scheme Code": scheme,
            "Row Type": row_type,
            "ISIN": isin,
            "Name of the Instrument": instrument,
            "Column": column,
            "Master Value": master_value,
            "SEBI Value": sebi_value,
            "Status": status,
            "SEBI Excel Row": excel_row,
        }
    )


def highlight_cell(ws, row, col, fill):
    if row and col:
        ws.cell(row, col).fill = fill



def normalize_key(value):
    return " ".join(clean_text(value).casefold().split())


def normalize_security_type(value):
    return " ".join(clean_text(value).casefold().split())


def is_blank_cell(value):
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return str(value).strip() == ""


def is_note_or_non_data_label(label):
    label_lower = clean_text(label).casefold()
    if not label_lower:
        return True
    if label_lower in {"sub total", "total", "grand total", "net receivables / (payables)"}:
        return True
    if label_lower.startswith("$"):
        return True
    if label_lower.startswith("^"):
        return True
    if "portfolio does not include" in label_lower:
        return True
    if "net assets includes" in label_lower:
        return True
    if "disclosure in derivatives" in label_lower:
        return True
    return False


def row_values_from_sebi(ws, row, col_map):
    rec = {"Excel Row": row}
    for col in SEBI_COLUMNS:
        rec[col] = ws.cell(row, col_map[col]).value
    return rec


def get_remaining_effective_section(label, current_section):
    """Map nested SEBI labels to the actual section to be compared."""
    label_clean = clean_text(label)

    # This label appears under Foreign Securities in equity schemes.
    # For comparison, treat rows below it as Foreign Securities.
    if label_clean == "(a) Listed / awaiting listing on Stock Exchanges":
        if current_section == "Foreign Securities and/or overseas ETF(s)":
            return "Foreign Securities and/or overseas ETF(s)"
        return label_clean

    return label_clean


def read_remaining_equity_sections(ws, first_total_row, col_map):
    """
    Read rows after the first Total for equity schemes.

    Cash and Bank / Margin (Future and Options) are single-line sections with values in
    Market/Fair Value, so their own row is captured.

    Other sections are captured until Sub Total / Total / next section.
    """
    name_col = col_map["Name of the Instrument"]
    current_section = None
    section_records = {}

    for row in range(first_total_row + 1, ws.max_row + 1):
        label = clean_text(ws.cell(row, name_col).value)
        if not label:
            continue

        label_lower = label.casefold()

        if label_lower in {"grand total", "net receivables / (payables)"}:
            current_section = None
            continue

        # New section/heading row.
        if label in REMAINING_EQUITY_SECTIONS:
            effective_section = get_remaining_effective_section(label, current_section)
            current_section = effective_section

            # Cash and Margin are value rows themselves. Capture their own row.
            if label in TRIAL_BALANCE_ACCOUNT_CODES:
                section_records.setdefault(label, []).append(row_values_from_sebi(ws, row, col_map))
            continue

        if label_lower in {"sub total", "total"}:
            # Do not compare Sub Total / Total for remaining sections yet.
            # Existing first-section subtotal/total logic is unchanged.
            continue

        if is_note_or_non_data_label(label):
            continue

        if current_section in REMAINING_SECTION_SECURITY_TYPES:
            rec = row_values_from_sebi(ws, row, col_map)
            section_records.setdefault(current_section, []).append(rec)

    return section_records


def source_quantity_for_remaining_compare(value):
    """For remaining sections, keep Quantity comparison same as existing testone.py logic.
    Do NOT divide Quantity by 100000.
    """
    dec_value = to_decimal(value)
    if dec_value is None:
        return None
    return round_to_2(dec_value)


def sebi_quantity_for_remaining_compare(value):
    dec_value = to_decimal(value)
    if dec_value is None:
        return None
    return round_to_2(dec_value)


def market_value_for_compare(value):
    dec_value = to_decimal(value)
    if dec_value is None:
        return None
    return round_to_2(dec_value)


def converted_market_value_from_source(value):
    """Convert source raw amount to SEBI lacs value."""
    dec_value = to_decimal(value)
    if dec_value is None:
        return None
    return round_to_2(dec_value / Decimal(str(MARKET_VALUE_DIVISOR)))


def same_value_remaining(master_value, sebi_value, column_name):
    """Comparison rules for remaining equity sections only."""
    if column_name == "Quantity":
        mv = source_quantity_for_remaining_compare(master_value)
        sv = sebi_quantity_for_remaining_compare(sebi_value)
        if mv is None and sv is None:
            return True
        if mv is None or sv is None:
            return False
        return mv == sv

    # Keep % logic same as existing equity logic for non-cash/non-margin sections.
    if column_name == "% to Net Assets":
        return same_value(master_value, sebi_value, column_name)

    return same_value(master_value, sebi_value, column_name)


def display_value_remaining(value, source, column_name):
    """Write comparison values in Index for remaining sections."""
    if column_name == "Quantity":
        if source == "MASTER":
            val = source_quantity_for_remaining_compare(value)
        else:
            val = sebi_quantity_for_remaining_compare(value)
        return "" if val is None else format(val, "f")

    if column_name == "Market/Fair Value\n(Rs.in Lacs)":
        val = market_value_for_compare(value)
        return "" if val is None else format(val, "f")

    return value


def prepare_master_remaining_table(master_df, scheme_code, section_name):
    allowed_types = REMAINING_SECTION_SECURITY_TYPES.get(section_name)
    if not allowed_types:
        return [], {}, {}

    allowed_norm = {normalize_security_type(x) for x in allowed_types}
    table = master_df[
        (master_df["Client Code"] == scheme_code)
        & (master_df["Security Type Name"].apply(normalize_security_type).isin(allowed_norm))
    ].copy()

    if table.empty:
        return [], {}, {}

    records = []
    for _, row in table.iterrows():
        name = row.get("Issuer Name")
        if is_blank_cell(name):
            name = row.get("Security Name")

        industry_or_rating = row.get("Industry")
        if is_blank_cell(industry_or_rating):
            industry_or_rating = row.get("Rating")

        rec = {
            "Name of the Instrument": name,
            "ISIN": row.get("ISIN"),
            "Industry^ / Rating": industry_or_rating,
            "Quantity": row.get("Quantity"),
            "Market/Fair Value\n(Rs.in Lacs)": converted_market_value_from_source(row.get("Total Market Value (Rs.)")),
            "% to Net Assets": row.get("% to Net assests"),
        }
        records.append(rec)

    by_isin = {normalize_isin(r.get("ISIN")): r for r in records if normalize_isin(r.get("ISIN"))}
    by_name = {normalize_key(r.get("Name of the Instrument")): r for r in records if normalize_key(r.get("Name of the Instrument"))}
    return records, by_isin, by_name


def prepare_sebi_remaining_lookup(records):
    by_isin = {normalize_isin(r.get("ISIN")): r for r in records if normalize_isin(r.get("ISIN"))}
    by_name = {normalize_key(r.get("Name of the Instrument")): r for r in records if normalize_key(r.get("Name of the Instrument"))}
    return by_isin, by_name



def compare_remaining_master_section(ws, scheme_code, section_name, sebi_records, master_df, all_discrepancies, col_map):
    master_records, master_by_isin, master_by_name = prepare_master_remaining_table(master_df, scheme_code, section_name)

    matched_master_ids = set()
    matched_sebi_ids = set()

    # Match each SEBI row with Master.
    # Priority 1: ISIN when SEBI has ISIN.
    # Priority 2: Name, useful for TREPS where SEBI has blank ISIN but Master may have an ISIN.
    for sebi_row in sebi_records:
        sebi_isin = normalize_isin(sebi_row.get("ISIN"))
        sebi_name_key = normalize_key(sebi_row.get("Name of the Instrument"))

        master_row = None
        matched_isin = sebi_isin

        if sebi_isin:
            master_row = master_by_isin.get(sebi_isin)

        if master_row is None and sebi_name_key:
            master_row = master_by_name.get(sebi_name_key)
            if master_row is not None:
                matched_isin = normalize_isin(master_row.get("ISIN"))

        if master_row is None:
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                f"Remaining Section - {section_name}",
                sebi_isin,
                sebi_row.get("Name of the Instrument"),
                "Entire Row",
                None,
                "Present",
                "Master workbook does not have this entire row",
                sebi_row.get("Excel Row"),
            )
            continue

        matched_master_ids.add(id(master_row))
        matched_sebi_ids.add(id(sebi_row))
        compare_remaining_row_values(ws, scheme_code, section_name, matched_isin, master_row, sebi_row, all_discrepancies, col_map)

    # Add Master rows that were not found in SEBI.
    for master_row in master_records:
        if id(master_row) in matched_master_ids:
            continue

        add_discrepancy(
            all_discrepancies,
            scheme_code,
            f"Remaining Section - {section_name}",
            normalize_isin(master_row.get("ISIN")),
            master_row.get("Name of the Instrument"),
            "Entire Row",
            "Present",
            None,
            "SEBI sheet does not have this entire row",
            None,
        )


def should_compare_remaining_column(section_name, col, master_row, sebi_row):
    # Some remaining-section rows, for example TREPS, do not carry Industry/Quantity in SEBI.
    # In that case, do not create false mismatches for non-applicable blank SEBI fields.
    if col in {"Industry^ / Rating", "Quantity"} and is_blank_cell(sebi_row.get(col)):
        return False
    return True


def compare_remaining_row_values(ws, scheme_code, section_name, isin, master_row, sebi_row, all_discrepancies, col_map):
    for col in REMAINING_COMPARE_COLS:
        if not should_compare_remaining_column(section_name, col, master_row, sebi_row):
            continue

        if not same_value_remaining(master_row.get(col), sebi_row.get(col), col):
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                f"Remaining Section - {section_name}",
                isin,
                master_row.get("Name of the Instrument") or sebi_row.get("Name of the Instrument"),
                col,
                display_value_remaining(master_row.get(col), "MASTER", col),
                display_value_remaining(sebi_row.get(col), "SEBI", col),
                "Value mismatch",
                sebi_row.get("Excel Row"),
            )
            highlight_cell(ws, sebi_row.get("Excel Row"), col_map.get(col), HIGHLIGHT_FILL)


def treps_master_type_key(value):
    """Return canonical TREPS type from Master Security Type Name.

    normalize_security_type() returns lowercase text, so compare with normalized
    lowercase labels and return one canonical uppercase key used by the TREPS
    matching logic. This prevents correct CBL/REPO rows from being missed due
    to case or the Master typo: LENDIN vs LENDING.
    """
    security_type = normalize_security_type(value)

    if security_type == normalize_security_type("REPO"):
        return "REPO"

    if security_type in {
        normalize_security_type("COLLATERALISED BORROWING AND LENDING"),
        normalize_security_type("COLLATERALISED BORROWING AND LENDIN"),
    }:
        return "COLLATERALISED BORROWING AND LENDING"

    return ""


def treps_sebi_type_key(value):
    name = normalize_key(value)
    if name == normalize_key("Reverse Repo"):
        return "REPO"
    if name == normalize_key("Clearing Corporation of India Limited"):
        return "COLLATERALISED BORROWING AND LENDING"
    return ""


def treps_compare_key(row, source):
    market_value = market_value_for_compare(row.get("Market/Fair Value\n(Rs.in Lacs)"))
    percent_value = percent_value_for_compare(row.get("% to Net Assets"), source)
    percent_value = round_to_2(percent_value) if percent_value is not None else None
    return market_value, percent_value


def format_decimal_for_index(value):
    if value is None:
        return ""
    return format(value, "f")


def prepare_master_treps_records(master_df, scheme_code):
    allowed_types = {"REPO", "COLLATERALISED BORROWING AND LENDING", "COLLATERALISED BORROWING AND LENDIN"}
    table = master_df[master_df["Client Code"] == scheme_code].copy()

    records = []
    for _, row in table.iterrows():
        type_key = treps_master_type_key(row.get("Security Type Name"))
        if type_key not in allowed_types:
            continue

        rec = {
            "TREPS Type": type_key,
            "Name of the Instrument": "Reverse Repo" if type_key == "REPO" else "Clearing Corporation of India Limited",
            "ISIN": row.get("ISIN"),
            "Market/Fair Value\n(Rs.in Lacs)": converted_market_value_from_source(row.get("Total Market Value (Rs.)")),
            "% to Net Assets": row.get("% to Net assests"),
        }
        records.append(rec)

    return records


def compare_treps_reverse_repo_section(ws, scheme_code, sebi_records, master_df, all_discrepancies, col_map):
    """Special TREPS comparison.

    Do not compare instrument name for this section.
    Match by: type + processed Market/Fair Value + processed % to Net Assets.
    Type mapping:
      Reverse Repo -> Master Security Type Name REPO
      Clearing Corporation of India Limited -> Master Security Type Name COLLATERALISED BORROWING AND LENDING
    """
    master_records = prepare_master_treps_records(master_df, scheme_code)

    master_buckets = defaultdict(list)
    for rec in master_records:
        key = (rec.get("TREPS Type"),) + treps_compare_key(rec, "MASTER")
        master_buckets[key].append(rec)

    matched_master_ids = set()

    for sebi_row in sebi_records:
        sebi_type = treps_sebi_type_key(sebi_row.get("Name of the Instrument"))
        if not sebi_type:
            continue

        key = (sebi_type,) + treps_compare_key(sebi_row, "SEBI")
        available_master_rows = [r for r in master_buckets.get(key, []) if id(r) not in matched_master_ids]

        if available_master_rows:
            matched_master_ids.add(id(available_master_rows[0]))
            continue

        market_value, percent_value = key[1], key[2]
        add_discrepancy(
            all_discrepancies,
            scheme_code,
            "Remaining Section - TREPS / Reverse Repo",
            normalize_isin(sebi_row.get("ISIN")),
            sebi_row.get("Name of the Instrument"),
            "Market/Fair Value + % to Net Assets",
            "No matching value combination in Master",
            f"Market/Fair Value={format_decimal_for_index(market_value)}, % to Net Assets={format_decimal_for_index(percent_value)}%",
            "Value combination mismatch",
            sebi_row.get("Excel Row"),
        )
        highlight_cell(ws, sebi_row.get("Excel Row"), col_map.get("Market/Fair Value\n(Rs.in Lacs)"), HIGHLIGHT_FILL)
        highlight_cell(ws, sebi_row.get("Excel Row"), col_map.get("% to Net Assets"), HIGHLIGHT_FILL)

    for master_row in master_records:
        if id(master_row) in matched_master_ids:
            continue

        market_value, percent_value = treps_compare_key(master_row, "MASTER")
        add_discrepancy(
            all_discrepancies,
            scheme_code,
            "Remaining Section - TREPS / Reverse Repo",
            normalize_isin(master_row.get("ISIN")),
            master_row.get("Name of the Instrument"),
            "Market/Fair Value + % to Net Assets",
            f"Market/Fair Value={format_decimal_for_index(market_value)}, % to Net Assets={format_decimal_for_index(percent_value)}%",
            None,
            "SEBI sheet does not have this value combination",
            None,
        )


def load_trial_balance_df():
    if not os.path.exists(TRIAL_BALANCE_FILE):
        return None
    return pd.read_excel(TRIAL_BALANCE_FILE, sheet_name=TRIAL_BALANCE_SHEET)


def trial_balance_market_value(trial_df, scheme_code, section_name):
    if trial_df is None:
        return None

    account_code = TRIAL_BALANCE_ACCOUNT_CODES[section_name]
    df = trial_df.copy()
    df["Client Code"] = df["Client Code"].astype(str).str.strip()
    df["Account Code"] = df["Account Code"].astype(str).str.strip()

    selected = df[
        (df["Client Code"] == scheme_code)
        & (df["Account Code"] == str(account_code))
    ]

    if selected.empty:
        return None

    total_opening_balance = Decimal("0")
    found_value = False
    for value in selected["Opening Balance"]:
        dec_value = to_decimal(value)
        if dec_value is not None:
            total_opening_balance += dec_value
            found_value = True

    if not found_value:
        return None

    return round_to_2(total_opening_balance / Decimal(str(MARKET_VALUE_DIVISOR)))


def compare_trial_balance_section(ws, scheme_code, section_name, sebi_records, trial_df, all_discrepancies, col_map):
    market_col = "Market/Fair Value\n(Rs.in Lacs)"

    if not sebi_records:
        return

    sebi_row = sebi_records[0]
    sebi_value = market_value_for_compare(sebi_row.get(market_col))
    trial_value = trial_balance_market_value(trial_df, scheme_code, section_name)

    if trial_df is None:
        add_discrepancy(
            all_discrepancies,
            scheme_code,
            f"Remaining Section - {section_name}",
            "",
            section_name,
            market_col,
            None,
            sebi_value,
            f"Trial Balance workbook not found: {TRIAL_BALANCE_FILE}",
            sebi_row.get("Excel Row"),
        )
        return

    if trial_value is None:
        add_discrepancy(
            all_discrepancies,
            scheme_code,
            f"Remaining Section - {section_name}",
            "",
            section_name,
            market_col,
            None,
            sebi_value,
            "Trial Balance does not have matching Client Code + Account Code row",
            sebi_row.get("Excel Row"),
        )
        return

    if trial_value != sebi_value:
        add_discrepancy(
            all_discrepancies,
            scheme_code,
            f"Remaining Section - {section_name}",
            "",
            section_name,
            market_col,
            trial_value,
            sebi_value,
            "Value mismatch",
            sebi_row.get("Excel Row"),
        )
        highlight_cell(ws, sebi_row.get("Excel Row"), col_map.get(market_col), HIGHLIGHT_FILL)

    # NOTE: % to Net Assets for Cash and Bank / Margin is intentionally not checked here.
    # It will be implemented later after the exact rule is provided.


def compare_remaining_equity_sections(ws, scheme_code, master_df, trial_df, all_discrepancies, sebi):
    if not sebi.get("total"):
        return

    col_map = sebi["col_map"]
    first_total_row = sebi["total"].get("Excel Row")
    section_records = read_remaining_equity_sections(ws, first_total_row, col_map)

    for section_name, records in section_records.items():
        if section_name == "TREPS / Reverse Repo":
            compare_treps_reverse_repo_section(ws, scheme_code, records, master_df, all_discrepancies, col_map)
        elif section_name in TRIAL_BALANCE_ACCOUNT_CODES:
            compare_trial_balance_section(ws, scheme_code, section_name, records, trial_df, all_discrepancies, col_map)
        elif section_name in REMAINING_SECTION_SECURITY_TYPES:
            compare_remaining_master_section(ws, scheme_code, section_name, records, master_df, all_discrepancies, col_map)


def compare_scheme(ws, scheme_code, master_df, all_discrepancies, trial_df=None):
    sebi = read_sebi_first_section(ws)
    master_table, master_records, master_by_isin = read_master_equity(master_df, scheme_code)
    sebi_by_isin = sebi["row_by_isin"]
    col_map = sebi["col_map"]

    # 1. Security row comparison by ISIN.
    all_isins = sorted(set(master_by_isin.keys()) | set(sebi_by_isin.keys()))
    for isin in all_isins:
        master_row = master_by_isin.get(isin)
        sebi_row = sebi_by_isin.get(isin)

        if master_row is None:
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                "Security Row",
                isin,
                sebi_row.get("Name of the Instrument"),
                "Entire Row",
                None,
                "Present",
                "Master workbook does not have this entire row",
                sebi_row.get("Excel Row"),
            )
            # Existing first-section behavior kept unchanged from your latest code.
            for col in SEBI_COLUMNS:
                highlight_cell(ws, sebi_row.get("Excel Row"), col_map.get(col), HIGHLIGHT_FILL)
            continue

        if sebi_row is None:
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                "Security Row",
                isin,
                master_row.get("Name of the Instrument"),
                "Entire Row",
                "Present",
                None,
                "SEBI sheet does not have this entire row",
                None,
            )
            continue

        for col in COMPARE_COLS:
            if not same_value(master_row.get(col), sebi_row.get(col), col):
                add_discrepancy(
                    all_discrepancies,
                    scheme_code,
                    "Security Row",
                    isin,
                    master_row.get("Name of the Instrument") or sebi_row.get("Name of the Instrument"),
                    col,
                    master_row.get(col),
                    sebi_row.get(col),
                    "Value mismatch",
                    sebi_row.get("Excel Row"),
                )
                highlight_cell(ws, sebi_row.get("Excel Row"), col_map.get(col), HIGHLIGHT_FILL)

    # 2. First Sub Total and first Total comparison.
    master_market_total = round(pd.to_numeric(master_table["Market/Fair Value\n(Rs.in Lacs)"], errors="coerce").fillna(0).sum(), 2)
    master_percent_total = pd.to_numeric(master_table["% to Net Assets"], errors="coerce").fillna(0).sum()

    for row_label, sebi_total_row in [("Sub Total", sebi["subtotal"]), ("Total", sebi["total"])] :
        if not sebi_total_row:
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                row_label,
                "",
                row_label,
                "Entire Row",
                "Present",
                None,
                f"SEBI sheet does not have first {row_label} row",
                None,
            )
            continue

        row_no = sebi_total_row.get("Excel Row")
        market_col = "Market/Fair Value\n(Rs.in Lacs)"
        pct_col = "% to Net Assets"

        if not same_value(master_market_total, sebi_total_row.get(market_col), market_col):
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                row_label,
                "",
                row_label,
                market_col,
                master_market_total,
                sebi_total_row.get(market_col),
                "Value mismatch",
                row_no,
            )
            highlight_cell(ws, row_no, col_map.get(market_col), TOTAL_FILL)

        if not same_value(master_percent_total, sebi_total_row.get(pct_col), pct_col):
            add_discrepancy(
                all_discrepancies,
                scheme_code,
                row_label,
                "",
                row_label,
                pct_col,
                master_percent_total,
                sebi_total_row.get(pct_col),
                "Value mismatch",
                row_no,
            )
            highlight_cell(ws, row_no, col_map.get(pct_col), TOTAL_FILL)

    # 3. Remaining equity sections after the first Total.
    compare_remaining_equity_sections(ws, scheme_code, master_df, trial_df, all_discrepancies, sebi)


def write_index_output(wb, discrepancies):
    if "Index" not in wb.sheetnames:
        ws = wb.create_sheet("Index", 0)
    else:
        ws = wb["Index"]

    # Clear old output from E onwards only.
    for row in range(1, ws.max_row + 1):
        for col in range(5, max(ws.max_column, 20) + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.font = copy(ws.cell(row, 1).font)
            cell.border = Border()
            cell.alignment = Alignment()

    title_cell = ws.cell(1, 5)
    title_cell.value = "Equity Reconciliation Output - First Total + Remaining Sections"
    title_cell.font = Font(bold=True, size=12)

    headers = [
        "Scheme Code",
        "Row Type",
        "ISIN",
        "Name of the Instrument",
        "Column",
        "Master Value",
        "SEBI Value",
        "Status",
        "SEBI Excel Row",
    ]

    start_row = 3
    start_col = 5
    for i, header in enumerate(headers, start=start_col):
        cell = ws.cell(start_row, i)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for r_idx, rec in enumerate(discrepancies, start=start_row + 1):
        for c_idx, header in enumerate(headers, start=start_col):
            cell = ws.cell(r_idx, c_idx)

            if rec.get("Column") == "% to Net Assets" and header == "Master Value":
                cell.value = format_percent_for_index(rec.get("Master Value"), "MASTER")

            elif rec.get("Column") == "% to Net Assets" and header == "SEBI Value":
                cell.value = format_percent_for_index(rec.get("SEBI Value"), "SEBI")

            else:
                cell.value = rec.get(header)

            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        5: 14, 6: 16, 7: 18, 8: 42, 9: 30, 10: 20, 11: 20, 12: 42, 13: 14
    }
    for col, width in widths.items():
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    ws.freeze_panes = "E4"
    ws.auto_filter.ref = f"E3:M{max(start_row + 1, start_row + len(discrepancies))}"


def main():
    source_xlsx = resolve_source_xlsx()

    wb = load_workbook(source_xlsx)
    master_df = pd.read_excel(MASTER_FILE)
    trial_df = load_trial_balance_df()

    discrepancies = []
    processed = []
    skipped = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "Index":
            continue
        if scheme_dict.get(sheet_name) != "EQUITY":
            continue

        ws = wb[sheet_name]
        try:
            compare_scheme(ws, sheet_name, master_df, discrepancies, trial_df)
            processed.append(sheet_name)
        except Exception as exc:
            skipped.append((sheet_name, str(exc)))
            add_discrepancy(
                discrepancies,
                sheet_name,
                "Sheet Error",
                "",
                "",
                "",
                "",
                "",
                f"Could not process sheet: {exc}",
                None,
            )

    write_index_output(wb, discrepancies)
    wb.save(OUTPUT_FILE)

    print(f"Processed equity schemes: {len(processed)}")
    print(f"Discrepancies written: {len(discrepancies)}")
    if skipped:
        print("Skipped sheets:")
        for item in skipped:
            print(item)
    print(f"Output file: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()